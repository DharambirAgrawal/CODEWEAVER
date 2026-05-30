#!/usr/bin/env python3
# src/execify/probes/probe_xlsx.py
# Run: python probe_xlsx.py <path_to_xlsx>
import sys
import json
try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

if len(sys.argv) < 2:
    print(json.dumps({"error": "Usage: probe_xlsx.py <path>"}))
    sys.exit(1)

path = sys.argv[1]
try:
    wb = load_workbook(path)
except Exception as e:
    print(json.dumps({"error": str(e)}))
    sys.exit(1)

ws = wb.active
data_rows = 0
for row in ws.iter_rows(min_row=2):
    if any(c.value is not None for c in row):
        data_rows += 1

col_count = ws.max_column or 0
sample = [ws.cell(row=i, column=2).value for i in range(2, min(6, (ws.max_row or 1) + 1))]
all_same = len(set(str(v) for v in sample if v is not None)) == 1 if sample else False

print(json.dumps({
    "data_rows": data_rows,
    "col_count": col_count,
    "all_same_values": all_same,
    "headers": [ws.cell(row=1, column=c).value for c in range(1, col_count + 1)],
}))
